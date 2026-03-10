"""Tools for parsing and normalizing academic data files.

Handles Excel, PDF, JSON and CSV file ingestion with LLM-assisted
structure interpretation and human-in-the-loop clarification.
"""

import json
from pathlib import Path
from typing import Dict, Any, List, Annotated

from langchain_core.tools import tool as langgraph_tool
from langgraph.prebuilt import InjectedState
from langgraph.types import interrupt


@langgraph_tool
def tool_extraer_datos_archivo(
    state: Annotated[dict, InjectedState] = None,  # type: ignore[assignment]
) -> str:
    """Extrae datos crudos de un archivo Excel, PDF, JSON o CSV subido por el usuario.

    Lee el archivo, extrae headers y filas de datos, y retorna una muestra
    para que el agente interprete la estructura.
    Soporta: .xlsx, .xls, .pdf, .json, .csv
    """
    state = state or {}
    file_path = state.get("file_path", "")

    if not file_path:
        return json.dumps({"tipo": "archivo", "error": "No hay archivo para procesar"})

    path = Path(file_path)
    if not path.exists():
        return json.dumps({"tipo": "archivo", "error": f"Archivo no encontrado: {file_path}"})

    ext = path.suffix.lower()

    try:
        if ext == ".json":
            return _parse_json(path)
        elif ext in (".xlsx", ".xls"):
            return _parse_excel(path)
        elif ext == ".csv":
            return _parse_csv(path)
        elif ext == ".pdf":
            return _parse_pdf(path)
        else:
            return json.dumps({"tipo": "archivo", "error": f"Formato no soportado: {ext}"})
    except Exception as e:
        return json.dumps({"tipo": "archivo", "error": f"Error al procesar archivo: {str(e)}"})


@langgraph_tool
def tool_normalizar_datos_examen(
    column_mapping: str,
    state: Annotated[dict, InjectedState] = None,  # type: ignore[assignment]
) -> str:
    """Normaliza datos crudos de un archivo al schema estándar de examen.

    Recibe un mapeo de columnas (generado por el agente tras analizar la estructura)
    y transforma los datos al formato interno.

    Args:
        column_mapping: JSON string con el mapeo de columnas. Ejemplo:
            {"dni_column": "Código", "nombre_column": "Nombre",
             "respuestas_columns": ["R1","R2","R3"],
             "nota_column": "Puntaje"}
    """
    state = state or {}
    file_path = state.get("file_path", "")

    if not file_path:
        return json.dumps({"tipo": "normalizacion", "error": "No hay archivo para normalizar"})

    try:
        mapping = json.loads(column_mapping)
    except json.JSONDecodeError:
        return json.dumps({"tipo": "normalizacion", "error": "column_mapping no es JSON válido"})

    path = Path(file_path)
    ext = path.suffix.lower()

    try:
        if ext in (".xlsx", ".xls"):
            rows, headers = _read_excel_full(path, mapping.get("sheet_name"))
        elif ext == ".csv":
            rows, headers = _read_csv_full(path)
        elif ext == ".json":
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return json.dumps({"tipo": "normalizacion", "datos": data, "mensaje": "JSON ya está en formato nativo"})
        else:
            return json.dumps({"tipo": "normalizacion", "error": f"Normalización no soportada para {ext}"})

        # Apply column mapping
        students = _apply_mapping(rows, headers, mapping)

        return json.dumps({
            "tipo": "normalizacion",
            "students_data": students,
            "total_normalizados": len(students),
            "mensaje": f"Normalizados {len(students)} registros de estudiantes"
        }, ensure_ascii=False)

    except Exception as e:
        return json.dumps({"tipo": "normalizacion", "error": f"Error al normalizar: {str(e)}"})


@langgraph_tool
def tool_solicitar_clarificacion(
    pregunta: str,
    contexto: str = "",
    opciones: str = "",
) -> str:
    """Solicita aclaración al usuario cuando los datos son ambiguos.

    Pausa la ejecución del agente y espera respuesta del usuario.
    Usar cuando:
    - La estructura del archivo no es clara
    - Hay columnas ambiguas que podrían ser diferentes campos
    - Faltan datos esperados
    - Se necesita confirmar un mapeo de datos

    Args:
        pregunta: Pregunta clara y específica para el usuario
        contexto: Información adicional para que el usuario entienda (muestra de datos, etc.)
        opciones: Lista de opciones separadas por '|' (vacío si es pregunta abierta)
    """
    opciones_list = [o.strip() for o in opciones.split("|") if o.strip()] if opciones else []

    respuesta = interrupt({
        "tipo": "clarificacion",
        "pregunta": pregunta,
        "contexto": contexto,
        "opciones": opciones_list,
    })

    return f"El usuario respondió: {respuesta}"


# === Internal parsing functions ===

def _parse_json(path: Path) -> str:
    """Parse a JSON file and return its structure."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Determine structure
    info: Dict[str, Any] = {"tipo": "archivo", "formato": "json"}

    if isinstance(data, dict):
        info["keys"] = list(data.keys())
        # Check if it matches exam schema
        if "examen" in data and "preguntas" in data:
            info["schema_detectado"] = "examen_nodaris"
            info["mensaje"] = "Archivo JSON con schema de examen Nodaris detectado"
            info["datos"] = data
        else:
            info["schema_detectado"] = "desconocido"
            info["muestra"] = {k: _summarize_value(v) for k, v in data.items()}
    elif isinstance(data, list):
        info["total_registros"] = len(data)
        if data:
            info["muestra_primer_registro"] = data[0] if isinstance(data[0], dict) else str(data[0])[:200]

    return json.dumps(info, ensure_ascii=False, default=str)


def _parse_excel(path: Path) -> str:
    """Parse an Excel file, return structure + auto-parsed data if possible."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    info: Dict[str, Any] = {
        "tipo": "archivo",
        "formato": "excel",
        "hojas": wb.sheetnames,
    }

    sheets_data = []
    active_headers_lower = None
    active_all_data_rows = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_active = (ws == wb.active)

        # Read all rows for active sheet (for auto-detection), preview for others
        if is_active:
            all_rows = list(ws.iter_rows(values_only=True))
        else:
            all_rows = list(ws.iter_rows(max_row=6, values_only=True))

        if not all_rows:
            sheets_data.append({"hoja": sheet_name, "vacia": True})
            continue

        headers = [str(c) if c is not None else "" for c in all_rows[0]]
        sample_rows = [
            [str(c) if c is not None else "" for c in row]
            for row in all_rows[1:6]
        ]

        sheet_info: Dict[str, Any] = {
            "hoja": sheet_name,
            "headers": headers,
            "muestra_filas": sample_rows,
        }

        if is_active:
            sheet_info["total_filas_estimado"] = len(all_rows) - 1
            active_headers_lower = [h.lower().strip() for h in headers]
            active_all_data_rows = all_rows[1:]
        else:
            sheet_info["total_filas_estimado"] = ws.max_row - 1 if ws.max_row else 0

        sheets_data.append(sheet_info)

    info["hojas_detalle"] = sheets_data
    wb.close()

    # Auto-detect data structure from active sheet
    if active_headers_lower and active_all_data_rows:
        try:
            from agent.nodes.validation import _try_build_from_tabular
            parsed = _try_build_from_tabular(active_headers_lower, active_all_data_rows)
            if parsed:
                exam_data, students = parsed
                info["auto_parsed"] = True
                info["datos"] = {"exam_data": exam_data, "students_data": students}
                info["mensaje"] = f"Excel auto-parseado: {len(students)} estudiantes detectados"
        except Exception:
            pass

    return json.dumps(info, ensure_ascii=False, default=str)


def _parse_csv(path: Path) -> str:
    """Parse a CSV file and return structure + auto-parsed data if possible."""
    import csv as csv_mod

    info: Dict[str, Any] = {
        "tipo": "archivo",
        "formato": "csv",
    }

    # Try common encodings
    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            with open(path, "r", encoding=encoding, newline="") as f:
                sniffer = csv_mod.Sniffer()
                sample = f.read(4096)
                try:
                    dialect = sniffer.sniff(sample)
                except csv_mod.Error:
                    dialect = csv_mod.excel  # type: ignore[assignment]
                f.seek(0)
                reader = csv_mod.reader(f, dialect)
                all_rows = list(reader)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        return json.dumps({"tipo": "archivo", "error": "No se pudo decodificar el archivo CSV"})

    if not all_rows:
        info["vacio"] = True
        return json.dumps(info, ensure_ascii=False)

    headers = all_rows[0]
    sample_rows = all_rows[1:6]  # up to 5 sample rows

    info["headers"] = headers
    info["muestra_filas"] = sample_rows
    info["total_filas_estimado"] = len(all_rows) - 1

    # Auto-detect data structure
    data_rows = all_rows[1:]
    if data_rows:
        headers_lower = [h.lower().strip() for h in headers]
        try:
            from agent.nodes.validation import _try_build_from_tabular
            parsed = _try_build_from_tabular(headers_lower, data_rows)
            if parsed:
                exam_data, students = parsed
                info["auto_parsed"] = True
                info["datos"] = {"exam_data": exam_data, "students_data": students}
                info["mensaje"] = f"CSV auto-parseado: {len(students)} estudiantes detectados"
        except Exception:
            pass

    return json.dumps(info, ensure_ascii=False, default=str)


def _read_csv_full(path: Path) -> tuple:
    """Read all data from a CSV file. Returns (data_rows, headers)."""
    import csv as csv_mod

    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            with open(path, "r", encoding=encoding, newline="") as f:
                sniffer = csv_mod.Sniffer()
                sample = f.read(4096)
                try:
                    dialect = sniffer.sniff(sample)
                except csv_mod.Error:
                    dialect = csv_mod.excel  # type: ignore[assignment]
                f.seek(0)
                reader = csv_mod.reader(f, dialect)
                all_rows = list(reader)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        return [], []

    if not all_rows:
        return [], []

    headers = all_rows[0]
    data_rows = all_rows[1:]
    return data_rows, headers


def _parse_pdf(path: Path) -> str:
    """Parse a PDF file and extract tables or text."""
    import pdfplumber

    info: Dict[str, Any] = {
        "tipo": "archivo",
        "formato": "pdf",
    }

    tables_found: List[Dict[str, Any]] = []
    text_content = []

    with pdfplumber.open(path) as pdf:
        info["total_paginas"] = len(pdf.pages)

        for i, page in enumerate(pdf.pages[:5]):  # Max 5 pages
            # Try to extract tables first
            tables = page.extract_tables()
            if tables:
                for t_idx, table in enumerate(tables):
                    if table and len(table) > 1:
                        tables_found.append({
                            "pagina": i + 1,
                            "tabla_idx": t_idx,
                            "headers": table[0],
                            "muestra_filas": table[1:4],
                            "total_filas": len(table) - 1,
                        })
            else:
                # Extract text as fallback
                text = page.extract_text()
                if text:
                    text_content.append({
                        "pagina": i + 1,
                        "texto": text[:500]
                    })

    if tables_found:
        info["tablas"] = tables_found
        info["tipo_contenido"] = "tablas"
    elif text_content:
        info["texto"] = text_content
        info["tipo_contenido"] = "texto_libre"
    else:
        info["tipo_contenido"] = "vacio"

    return json.dumps(info, ensure_ascii=False, default=str)


def _read_excel_full(path: Path, sheet_name: str = None) -> tuple:
    """Read all data from an Excel sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], []

    headers = [str(c) if c is not None else "" for c in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append([c for c in row])

    return data_rows, headers


def _apply_mapping(
    rows: List[list],
    headers: List[str],
    mapping: Dict[str, Any],
) -> List[Dict[str, Any]]:
    """Apply column mapping to transform raw rows into student records."""
    col_index = {h: i for i, h in enumerate(headers)}

    dni_col = mapping.get("dni_column", "")
    nombre_col = mapping.get("nombre_column", "")
    nota_col = mapping.get("nota_column", "")
    resp_cols = mapping.get("respuestas_columns", [])

    students = []
    for row in rows:
        student: Dict[str, Any] = {}

        if dni_col and dni_col in col_index:
            student["dni"] = str(row[col_index[dni_col]] or "")
        if nombre_col and nombre_col in col_index:
            student["nombre"] = str(row[col_index[nombre_col]] or "")
        if nota_col and nota_col in col_index:
            val = row[col_index[nota_col]]
            student["nota"] = float(val) if val is not None else 0
        if resp_cols:
            respuestas = []
            for rc in resp_cols:
                if rc in col_index:
                    val = row[col_index[rc]]
                    respuestas.append(str(val) if val is not None else "NR")
                else:
                    respuestas.append("NR")
            student["respuestas"] = respuestas

        if student:
            students.append(student)

    return students


def _summarize_value(v: Any) -> str:
    """Summarize a value for preview."""
    if isinstance(v, list):
        return f"list[{len(v)} items]"
    elif isinstance(v, dict):
        return f"dict[keys: {list(v.keys())[:5]}]"
    else:
        s = str(v)
        return s[:100] if len(s) > 100 else s
